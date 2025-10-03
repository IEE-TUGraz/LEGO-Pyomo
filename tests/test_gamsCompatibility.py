import shutil

from openpyxl import load_workbook

import InOutModule.Utilities
from InOutModule.CaseStudy import CaseStudy
from InOutModule.ExcelWriter import ExcelWriter
from InOutModule.printer import Printer
from LEGO.helpers.CompareModels import compareModels, ModelTypeForComparison

printer = Printer.getInstance()


def test_comparisonExampleAgainstGAMS(tmp_path):
    """
    Compares the GAMS result against the Pyomo result.
    :param tmp_path: Temporary path for the test (provided by pytest).
    :return: None
    """
    data_folder = "data/example"
    tmp_folder = tmp_path / "modified"

    # Copy everything so that Parameter files are also present
    shutil.copytree(data_folder, tmp_folder)

    # Adjust case study to have all inflows already as capacity factors for GAMS compatibility
    cs = CaseStudy(data_folder)
    cs.dPower_VRESProfiles = InOutModule.Utilities.inflowsToCapacityFactors(cs.dPower_Inflows, cs.dPower_VRES, cs.dPower_VRESProfiles)
    cs.dPower_Inflows = cs.dPower_Inflows[0:0]

    cs.dPower_Storage.drop("StorageHydro", inplace=True)  # Remove StorageHydro as it is implemented differently in GAMS version

    ew = ExcelWriter()
    ew.write_caseStudy(cs, tmp_folder)

    # Note: Storage is slightly different in GAMS version, so skip related variables/constraints
    mps_equal = compareModels(ModelTypeForComparison.DETERMINISTIC, tmp_folder, True,
                              ModelTypeForComparison.GAMS, tmp_folder, True,
                              coefficients_skip_model1=["vCurtailment", "vStorageSpillage", "vStIntraRes", "vStInterRes"], constraint_skip_model1=["eStInterRes", "eStMaxInterRes"],
                              coefficients_skip_model2=["vStIntraRes", "vStInterRes"], constraint_skip_model2=["eStInterRes", "eStMaxInterRes"],
                              tmp_folder_path=tmp_path, print_additional_information=True)

    assert mps_equal


def test_comparisonExampleSOCPAgainstGAMS(tmp_path):
    """
    Compares the GAMS result against the Pyomo result for the SOCP example.
    :param tmp_path: Temporary path for the test (provided by pytest).
    :return: None
    """
    data_folder = "data/example"

    # Copy the data folder to a temporary path
    tmp_path_originalData = str(tmp_path / "originalData")
    shutil.copytree(data_folder, tmp_path_originalData)

    # Switch solver (to enable quadratic constraints)
    workbook = load_workbook(filename=tmp_path_originalData + "/Global_Parameters.xlsx")
    sheet = workbook.active
    sheet["C5"] = "gurobi"  # Set solver
    workbook.save(filename=tmp_path_originalData + "/Global_Parameters.xlsx")

    # Activate SOCP in Power Parameters
    workbook = load_workbook(filename=tmp_path_originalData + "/Power_Parameters.xlsx")
    sheet = workbook.active
    sheet["C34"] = "Yes"  # Enable SOCP
    workbook.save(filename=tmp_path_originalData + "/Power_Parameters.xlsx")

    # Use SOCP for all lines
    workbook = load_workbook(filename=tmp_path_originalData + "/Power_Network.xlsx")
    sheet = workbook.active
    for i in range(8, sheet.max_row + 1):
        sheet[f"O{i}"] = "SOCP"  # Set all lines to use SOCP
    workbook.save(filename=tmp_path_originalData + "/Power_Network.xlsx")

    # Adjust case study to have all inflows already as capacity factors for GAMS compatibility
    cs = CaseStudy(tmp_path_originalData)
    cs.dPower_VRESProfiles = InOutModule.Utilities.inflowsToCapacityFactors(cs.dPower_Inflows, cs.dPower_VRES, cs.dPower_VRESProfiles)
    cs.dPower_Inflows = cs.dPower_Inflows[0:0]

    cs.dPower_Storage.drop("StorageHydro", inplace=True)  # Remove StorageHydro as it is implemented differently in GAMS version

    ew = ExcelWriter()
    ew.write_caseStudy(cs, tmp_path_originalData)

    mps_equal = compareModels(ModelTypeForComparison.DETERMINISTIC, tmp_path_originalData, True,
                              ModelTypeForComparison.GAMS, tmp_path_originalData, True,
                              coefficients_skip_model1=["vCurtailment", "vStorageSpillage", "vStIntraRes", "vStInterRes"], constraint_skip_model1=["eStInterRes", "eStMaxInterRes"],
                              coefficients_skip_model2=["vStIntraRes", "vStInterRes"], constraint_skip_model2=["eStInterRes", "eStMaxInterRes"],
                              tmp_folder_path=tmp_path, print_additional_information=True)

    assert mps_equal
